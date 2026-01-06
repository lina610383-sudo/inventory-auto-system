import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import io
import re

# 頁面配置
st.set_page_config(page_title="領用單自動化系統_完整版", layout="wide")
st.title("🚀 領用單流程自動化系統 (含領用人資訊填寫)")

def get_col_idx_by_header(ws, header_row_idx, target_field_key):
    """
    動態偵測：根據預設的關鍵字組，在指定標題列搜尋對應的欄位索引 (1-based)
    針對「名詞差異」進行強化模糊匹配
    """
    synonyms = {
        "Vendor": ["VENDOR", "SUPPLIER", "廠商", "供應商", "MFR", "Manufacturer"],
        "Description": ["DESCRIPTION", "品名", "描述", "零件名稱", "SPEC", "規格"],
        "HP PN": ["HP PN", "HPPN", "HP料號", "CUSTOMER PN", "客戶料號"],
        "IEC PN": ["IEC PN", "IECPN", "IEC料號", "INTERNAL PN", "內部料號", "料號"],
        "Unit": ["UNIT", "單位", "UOM"],
        "No": ["NO", "NO.", "項次", "序號", "INDEX"]
    }
    
    search_keywords = synonyms.get(target_field_key, [target_field_key])
    
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row_idx, column=col).value
        if val:
            cell_text = str(val).strip().upper()
            # 1. 優先完全匹配
            if any(k.upper() == cell_text for k in search_keywords):
                return col
            # 2. 包含匹配 (處理標題帶有空格或括號的情況)
            if any(k.upper() in cell_text for k in search_keywords):
                return col
    return None

def fill_personnel_info(ws, personnel_data):
    """
    在模板中搜尋「領用人」、「工號」、「部門」等標籤，並在對應位置填寫
    """
    tags = {
        "領用人": personnel_data.get('name', ''),
        "姓名": personnel_data.get('name', ''),
        "工號": personnel_data.get('id', ''),
        "員工編號": personnel_data.get('id', ''),
        "部門": personnel_data.get('dept', '')
    }
    
    # 掃描前 10 列尋找表頭資訊標籤
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val and isinstance(cell_val, str):
                for tag, value in tags.items():
                    if tag in cell_val:
                        # 檢查右側或下方是否有空白格可填入
                        if not ws.cell(row=r, column=c+1).value:
                            ws.cell(row=r, column=c+1, value=value)
                        break

def get_source_data(row, field_key):
    """
    根據關鍵字從 DataFrame 的 Row 中抓取資料，解決資料源 A 表的命名差異
    """
    source_synonyms = {
        "Vendor": ["Vendor", "Supplier", "廠商", "供應商"],
        "Description": ["Description", "品名", "描述", "零件名稱", "Description/品名"],
        "HP PN": ["HP PN", "HPPN", "Customer PN", "客戶料號"],
        "IEC PN": ["IEC PN", "IECPN", "Internal PN", "內部料號", "料號"],
        "Unit": ["Unit", "單位"]
    }
    
    potential_keys = source_synonyms.get(field_key, [field_key])
    for k in potential_keys:
        if k in row and pd.notna(row[k]):
            return row[k]
    return ""

def process_excel(file):
    try:
        # 1. 載入原始活頁簿
        wb = openpyxl.load_workbook(file)
        sheet_names = wb.sheetnames
        
        # 尋找目標明細分頁 (未開單)
        pattern = r".*領用明細_(\d+).*\(未開單\)"
        matches = []
        for s in sheet_names:
            m = re.search(pattern, s)
            if m:
                matches.append((m.group(1), s))
        
        if not matches:
            st.error("❌ 找不到符合格式的分頁！")
            return None, None
        
        latest_date, target_sheet_name = sorted(matches, key=lambda x: x[0])[-1]
        st.info(f"📍 目標明細分頁：{target_sheet_name}")
        
        # 2. 讀取資料
        detail_df = pd.read_excel(file, sheet_name=target_sheet_name, header=1)
        
        if "掛帳人清單" not in sheet_names:
            st.error("❌ 找不到『掛帳人清單』分頁！")
            return None, None
            
        payer_df = pd.read_excel(file, sheet_name="掛帳人清單")
        payer_df.iloc[:, 0] = payer_df.iloc[:, 0].ffill() 
        
        payer_map = {}
        for _, row in payer_df.iterrows():
            name = str(row['領用人']).strip()
            unit_type = str(row.iloc[0]).strip().upper() 
            if name and name != 'nan':
                payer_map[name] = {
                    'type': "IEC" if "IEC" in unit_type else "ICC",
                    'id': str(row['掛帳人']).strip(),
                    'name': name,
                    'dept': str(row.get('部門', '')).strip()
                }

        # 3. 準備產出分頁
        output_ws_dict = {}
        current_row_dict = {} 
        for t in ['IEC', 'ICC']:
            tmpl_name = f"領用單格式範例 {t}"
            if tmpl_name in sheet_names:
                new_ws = wb.copy_worksheet(wb[tmpl_name])
                new_ws.title = f"{t}_領用單_{latest_date}"
                output_ws_dict[t] = new_ws
                current_row_dict[t] = 6 
            else:
                st.warning(f"⚠️ 檔案中缺少模板：『{tmpl_name}』")

        # 4. 定位與回填
        valid_person_cols = [c for c in detail_df.columns if str(c).strip() in payer_map]
        filled_count = 0
        fields_to_sync = ["No", "Vendor", "Description", "HP PN", "IEC PN", "Unit"]

        personnel_filled = {"IEC": False, "ICC": False}

        for index, row in detail_df.iterrows():
            item_pn = get_source_data(row, 'IEC PN')
            if not item_pn: continue
            
            unit_targets = set()
            for person in valid_person_cols:
                qty = row[person]
                if pd.notna(qty) and isinstance(qty, (int, float)) and qty > 0:
                    unit_targets.add(payer_map[str(person).strip()]['type'])

            for t in unit_targets:
                if t in output_ws_dict:
                    ws = output_ws_dict[t]
                    target_row = current_row_dict[t]
                    
                    # A. 填入人員資訊
                    if not personnel_filled[t]:
                        try:
                            first_person = next(p for p in valid_person_cols if payer_map[str(p).strip()]['type'] == t)
                            fill_personnel_info(ws, payer_map[str(first_person).strip()])
                            personnel_filled[t] = True
                        except StopIteration:
                            pass

                    # B. 填入料件基本資訊 (自動適應 A 表與 B 表名詞差異)
                    for field in fields_to_sync:
                        col_idx = get_col_idx_by_header(ws, 5, field)
                        if col_idx:
                            if field == "No":
                                ws.cell(row=target_row, column=col_idx, value=target_row - 5)
                            else:
                                val = get_source_data(row, field)
                                if val:
                                    ws.cell(row=target_row, column=col_idx, value=val)
                    
                    # C. 填入領用數量 (掛帳人工號對位)
                    for person in valid_person_cols:
                        person_name = str(person).strip()
                        info = payer_map[person_name]
                        if info['type'] == t:
                            qty = row[person]
                            if pd.notna(qty) and isinstance(qty, (int, float)) and qty > 0:
                                target_col = None
                                target_id = str(info['id']).strip().upper()
                                for c in range(1, ws.max_column + 1):
                                    h_val = ws.cell(row=5, column=c).value
                                    if h_val and str(h_val).strip().upper() == target_id:
                                        target_col = c
                                        break
                                if target_col:
                                    ws.cell(row=target_row, column=target_col, value=qty)
                                    filled_count += 1
                    
                    current_row_dict[t] += 1

        # 5. 輸出
        ws_orig = wb[target_sheet_name]
        ws_orig.title = target_sheet_name.replace("(未開單)", "(已開單)")
        
        if filled_count > 0:
            st.success(f"✅ 完成！已處理名詞差異並同步資料。")
        else:
            st.warning("⚠️ 處理完成，但未發現有效的領用數量。")

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue(), latest_date

    except Exception as e:
        st.error(f"❌ 發生錯誤：{str(e)}")
        return None, None

# UI
uploaded_file = st.file_uploader("📂 請上傳領用單 Excel 檔案", type=["xlsx"])
if uploaded_file:
    if st.button("✨ 執行全自動生成"):
        processed_data, date = process_excel(uploaded_file)
        if processed_data:
            st.download_button("📥 下載領用單結果", data=processed_data, file_name=f"領用單完整產出_{date}.xlsx")
